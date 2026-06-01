from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import List
import fitz
import pandas as pd
import re
import os
import shutil
from openpyxl import load_workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

VALID_DESCRIPTIONS = [
    "Service Fee",
    "Marketing services",
    "Lead Gen Fees",
    "Recovery for additional cashbacks given to customers by restaurants",
    "Packing Charges",
    "Delivery Charges",
    "Advertisement Fees",
    "Cancellation Charges",
    "Convenience Fee"
]

def clean_text(text):
    if text is None:
        return ""
    text = str(text).replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def clean_number(value):
    if value is None:
        return 0.0
    value = str(value).strip()
    value = value.replace(",", "").replace("INR", "").replace("Rs.", "")
    try:
        return float(value)
    except:
        return 0.0

def extract(pattern, text):
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return clean_text(match.group(1)) if match else ""

def process_swiggy_pdf(full_path, file_name):
    results = []

    doc = fitz.open(full_path)
    full_text = ""

    for page in doc:
        full_text += page.get_text()

    doc.close()

    text = clean_text(full_text)

    invoice_no = extract(r'Invoice Number\s*:\s*([A-Z0-9\-]+)', text)
    invoice_date = extract(r'Invoice Date\s*:\s*([0-9:\-\s]+)', text)

    service_period_full = extract(
        r'Service Period\s*:\s*(.*?)PO Number',
        text
    )

    service_period_parts = re.split(
        r'Address\s*:',
        service_period_full,
        flags=re.IGNORECASE
    )

    service_period = service_period_parts[0].strip()
    service_address = ""

    if len(service_period_parts) > 1:
        service_address = service_period_parts[1].strip()

    irn = extract(r'IRN\s*:\s*([A-Za-z0-9]+)', text)

    company_name = "Swiggy Limited"

    company_match = re.search(
        r'GSTIN\s*:\s*([A-Z0-9]+)',
        text,
        re.IGNORECASE
    )

    company_gstin = company_match.group(1).strip() if company_match else ""

    customer_name = extract(
        r'Legal Name\s*:\s*(.*?)Restaurant',
        text
    )

    restaurant_name = extract(
        r'Restaurant\s*/\s*Store Name\s*:\s*(.*?)State Code',
        text
    )

    restaurant_id = extract(
        r'Restaurant\s*/\s*Store ID\s*:\s*([0-9]+)',
        text
    )
    
    # Restaurant ID Fallback Modification
    if restaurant_id == "":
        restaurant_id = extract(
            r'Restaurant\s*/\s*Store Name\s*:\s*.*?\(([0-9]+)\)',
            text
        )

    gstin_matches = re.findall(
        r'GSTIN\s*:\s*([A-Z0-9]+)',
        text,
        re.IGNORECASE
    )

    customer_gstin = gstin_matches[1] if len(gstin_matches) >= 2 else ""

    grand_total = clean_number(
        extract(r'Grand Total\s*([0-9,\.]+)', text)
    )

    reimbursement_discount = clean_number(
        extract(
            r'Other Charges\s*-\s*Reimbursement\s*of\s*Discount\s*([0-9,\.]+)',
            text
        )
    )

    table_start = text.find("Sr.")

    if table_start == -1:
        return []

    table_text = text[table_start:]

    line_pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Za-z0-9 \-\&\/\(\)\.\,\%]+?)\s+'
        r'(\d{6})\s+'
        r'OTH\s+'
        r'(\d+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9\.]*)\s+'
        r'([0-9,\.\-]*)\s+'
        r'([0-9\.]*)\s+'
        r'([0-9,\.\-]*)\s+'
        r'([0-9\.]*)\s+'
        r'([0-9,\.\-]*)',
        re.IGNORECASE
    )

    matches = line_pattern.findall(table_text)

    if not matches:
        return []

    # Updated invoice_row mapping structure
    invoice_row = {
        "File Name": file_name,
        "Invoice No": invoice_no,
        "Invoice Date": invoice_date,
        "Service Period": service_period,
        "Service Address": service_address,
        "IRN": irn,
        "Company Name": company_name,
        "Company GSTIN": company_gstin,
        "Customer Name": customer_name,
        "Restaurant Name": restaurant_name,
        "Restaurant ID": restaurant_id,
        "Customer GSTIN": customer_gstin,
        "Unit Price": 0.0,
        "Base Amount": 0.0,
        "Discount": 0.0,
        "Taxable Value": 0.0,
        "CGST Rate": 0.0,
        "CGST Amount": 0.0,
        "SGST Rate": 0.0,
        "SGST Amount": 0.0,
        "IGST Rate": 0.0,
        "IGST Amount": 0.0,
        "Line Total": 0.0,
        "Other Charges - Reimbursement of Discount": reimbursement_discount,
        "Grand Total": grand_total,
        "Service Fee": 0.0,
        "Lead Gen Fees": 0.0,
        "Marketing services": 0.0
    }

    # Tax Rate Counters Initialization
    cgst_rate_count = 0
    sgst_rate_count = 0
    igst_rate_count = 0

    for row in matches:
        description = clean_text(row[1])

        # Full Row Parsing Implementation
        unit_price = clean_number(row[4])
        base_amount = clean_number(row[5])
        discount = clean_number(row[6])
        assessable_value = clean_number(row[7])
        cgst_rate = clean_number(row[8])
        cgst_amount = clean_number(row[9])
        sgst_rate = clean_number(row[10])
        sgst_amount = clean_number(row[11])
        igst_rate = clean_number(row[12])
        igst_amount = clean_number(row[13])

        line_total = round(
            assessable_value +
            cgst_amount +
            sgst_amount +
            igst_amount,
            2
        )
        
        final_description = ""

        for desc in VALID_DESCRIPTIONS:
            if desc.lower() in description.lower():
                final_description = desc
                break

        if not final_description:
            continue

        # Accumulating Totals
        invoice_row["Unit Price"] += unit_price
        invoice_row["Base Amount"] += base_amount
        invoice_row["Discount"] += discount
        invoice_row["Taxable Value"] += assessable_value
        invoice_row["CGST Amount"] += cgst_amount
        invoice_row["SGST Amount"] += sgst_amount
        invoice_row["IGST Amount"] += igst_amount
        invoice_row["Line Total"] += line_total

        # Tax Rate Accumulation Loop Logic
        if cgst_rate > 0:
            invoice_row["CGST Rate"] += cgst_rate
            cgst_rate_count += 1

        if sgst_rate > 0:
            invoice_row["SGST Rate"] += sgst_rate
            sgst_rate_count += 1

        if igst_rate > 0:
            invoice_row["IGST Rate"] += igst_rate
            igst_rate_count += 1

        if final_description not in invoice_row:
            invoice_row[final_description] = 0.0

        invoice_row[final_description] += assessable_value

    # Compute averaged final Tax Rates after the loop finishes
    if cgst_rate_count:
        invoice_row["CGST Rate"] = round(invoice_row["CGST Rate"] / cgst_rate_count, 2)

    if sgst_rate_count:
        invoice_row["SGST Rate"] = round(invoice_row["SGST Rate"] / sgst_rate_count, 2)

    if igst_rate_count:
        invoice_row["IGST Rate"] = round(invoice_row["IGST Rate"] / igst_rate_count, 2)

    results.append(invoice_row)
    return results

@app.get("/")
def home():
    return {"message": "Swiggy Invoice API Running"}

@app.post("/swiggy")
async def upload_swiggy(files: List[UploadFile] = File(...)):
    results = []

    for file in files:
        safe_filename = os.path.basename(file.filename)
        save_path = os.path.join(UPLOAD_FOLDER, safe_filename)

        with open(save_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        try:
            results.extend(
                process_swiggy_pdf(save_path, safe_filename)
            )
        except Exception as e:
            print(e)

    df = pd.DataFrame(results)

    if df.empty:
        df = pd.DataFrame([{"Message": "No Data Found"}])

    output_file = os.path.join(
        OUTPUT_FOLDER,
        "Swiggy_Invoice_Output.xlsx"
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # Apply Float/Int "0.00" Number Formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # Set Column Auto Width Bounds
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = width + 5

    wb.save(output_file)
    wb.close()

    return FileResponse(
        path=output_file,
        filename="Swiggy_Invoice_Output.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )