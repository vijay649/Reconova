
# from fastapi import FastAPI, UploadFile, File, HTTPException
# from fastapi.middleware.cors import CORSMiddleware
# from fastapi.responses import FileResponse
# from typing import List
# import os
# import shutil
# import pandas as pd
# from datetime import datetime

# app = FastAPI()

# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"],
# )

# UPLOAD_DIR = "uploads"
# OUTPUT_DIR = "outputs"

# os.makedirs(UPLOAD_DIR, exist_ok=True)
# os.makedirs(OUTPUT_DIR, exist_ok=True)


# @app.get("/")
# def home():
#     return {"message": "Backend Running Successfully"}


# @app.post("/upload")
# async def upload_files(files: List[UploadFile] = File(...)):

#     if len(files) == 0:
#         raise HTTPException(status_code=400, detail="No files uploaded")

#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

#     upload_folder = os.path.join(UPLOAD_DIR, timestamp)

#     os.makedirs(upload_folder, exist_ok=True)

#     data = []

#     for file in files:

#         if not file.filename.lower().endswith(".pdf"):
#             continue

#         file_path = os.path.join(upload_folder, file.filename)

#         with open(file_path, "wb") as buffer:
#             shutil.copyfileobj(file.file, buffer)

#         data.append({
#             "File Name": file.filename,
#             "Status": "Processed Successfully"
#         })

#     if len(data) == 0:
#         raise HTTPException(status_code=400, detail="No PDF files found")

#     df = pd.DataFrame(data)

#     excel_path = os.path.join(
#         OUTPUT_DIR,
#         f"Amazon_Invoice_Output_{timestamp}.xlsx"
#     )

#     df.to_excel(excel_path, index=False, engine="openpyxl")

#     return FileResponse(
#         path=excel_path,
#         filename=f"Amazon_Invoice_Output_{timestamp}.xlsx",
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )


# import pandas as pd
# import fitz
# import re
# import os
# import shutil

# from fastapi import FastAPI, UploadFile, File
# from fastapi.middleware.cors import CORSMiddleware
# from fastapi.responses import StreamingResponse
# from typing import List
# from io import BytesIO

# app = FastAPI()

# # =====================================
# # CORS
# # =====================================

# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"],
# )

# # =====================================
# # UPLOAD FOLDER
# # =====================================

# UPLOAD_DIR = "uploads"
# os.makedirs(UPLOAD_DIR, exist_ok=True)

# # =====================================
# # CLEAN TEXT
# # =====================================

# def clean_text(text):

#     if not text:
#         return ""

#     text = text.replace("\n", " ")
#     text = re.sub(r"\s+", " ", text)

#     return text.strip()

# # =====================================
# # CLEAN NUMBER
# # =====================================

# def clean_number(value):

#     if not value:
#         return 0.0

#     value = str(value)

#     value = value.replace("INR", "")
#     value = value.replace(",", "")
#     value = value.strip()

#     try:
#         return float(value)

#     except:
#         return 0.0

# # =====================================
# # EXTRACT REGEX
# # =====================================

# def extract(pattern, text):

#     match = re.search(
#         pattern,
#         text,
#         re.IGNORECASE | re.DOTALL
#     )

#     if match:
#         return clean_text(match.group(1))

#     return ""

# # =====================================
# # PROCESS PDF
# # =====================================

# def process_pdf(pdf_path):

#     results = []

#     file_name = os.path.basename(pdf_path)

#     doc = fitz.open(pdf_path)

#     full_text = ""

#     for page in doc:
#         full_text += page.get_text()

#     text = clean_text(full_text)

#     # =================================
#     # HEADER DATA
#     # =================================

#     company_name = extract(
#         r"(Amazon Seller Services Private Limited|Amazon Retail India Private Limited)",
#         text
#     )

#     invoice_number = extract(
#         r"Invoice Number:\s*([A-Z0-9\-]+)",
#         text
#     )

#     invoice_date = extract(
#         r"Invoice Date:\s*([0-9\/]+)",
#         text
#     )

#     gstin = extract(
#         r"GST Tax Registration No:\s*([A-Z0-9]+)",
#         text
#     )

#     customer_name = extract(
#         r"Bill to\s*Name:\s*(.*?)\s*Address:",
#         text
#     )

#     state = extract(
#         r"Place of Supply:\s*([A-Z ]+)",
#         text
#     )

#     service_code = extract(
#         r"1\.\s*(\d{6})",
#         text
#     )

#     description = extract(
#         r"\d{6}\s+(.*?)\s+INR",
#         text
#     )

#     fee_amount = extract(
#         r"Storage Fee INR\s*([0-9\.]+)",
#         text
#     )

#     sgst_rate = extract(
#         r"SGST\s*([0-9\.]+)%",
#         text
#     )

#     sgst_amount = extract(
#         r"SGST\s*[0-9\.]+%\s*INR\s*([0-9\.]+)",
#         text
#     )

#     cgst_rate = extract(
#         r"CGST\s*([0-9\.]+)%",
#         text
#     )

#     cgst_amount = extract(
#         r"CGST\s*[0-9\.]+%\s*INR\s*([0-9\.]+)",
#         text
#     )

#     invoice_total = extract(
#         r"Total Invoice amount INR\s*([0-9\.]+)",
#         text
#     )

#     # =================================
#     # FINAL ROW
#     # =================================

#     results.append({
#         "File Name": file_name,
#         "Company Name": company_name,
#         "Invoice Number": invoice_number,
#         "Invoice Date": invoice_date,
#         "GSTIN": gstin,
#         "Customer Name": customer_name,
#         "State": state,
#         "Service Code": service_code,
#         "Description": description,
#         "Fee Amount": clean_number(fee_amount),
#         "SGST Rate": clean_number(sgst_rate),
#         "SGST Amount": clean_number(sgst_amount),
#         "CGST Rate": clean_number(cgst_rate),
#         "CGST Amount": clean_number(cgst_amount),
#         "Invoice Total": clean_number(invoice_total)
#     })

#     return results

# # =====================================
# # API
# # =====================================

# @app.post("/upload")
# async def upload_files(files: List[UploadFile] = File(...)):

#     all_results = []

#     for file in files:

#         if not file.filename.lower().endswith(".pdf"):
#             continue

#         save_path = os.path.join(
#             UPLOAD_DIR,
#             file.filename
#         )

#         with open(save_path, "wb") as buffer:
#             shutil.copyfileobj(file.file, buffer)

#         extracted = process_pdf(save_path)

#         all_results.extend(extracted)

#     # =================================
#     # DATAFRAME
#     # =================================

#     df = pd.DataFrame(all_results)

#     # =================================
#     # CREATE EXCEL
#     # =================================

#     excel_buffer = BytesIO()

#     with pd.ExcelWriter(
#         excel_buffer,
#         engine="openpyxl"
#     ) as writer:

#         df.to_excel(
#             writer,
#             index=False
#         )

#         worksheet = writer.sheets["Sheet1"]

#         # Auto width
#         for column in worksheet.columns:

#             max_length = 0

#             column_letter = column[0].column_letter

#             for cell in column:

#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))

#                 except:
#                     pass

#             adjusted_width = max_length + 5

#             worksheet.column_dimensions[
#                 column_letter
#             ].width = adjusted_width

#     excel_buffer.seek(0)

#     # =================================
#     # RETURN EXCEL
#     # =================================

#     return StreamingResponse(
#         excel_buffer,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Content-Disposition":
#             "attachment; filename=Amazon_Invoice_Output.xlsx"
#         }
#     )

# # =====================================
# # ROOT API
# # =====================================

# @app.get("/")
# def home():

#     return {
#         "message": "Amazon Invoice Backend Running Successfully"
#     }
    
# import fitz
# import pandas as pd
# import re
# import os

# from tkinter import Tk
# from tkinter.filedialog import askdirectory
# from tkinter import messagebox

# # =====================================
# # FOLDER PICKER GUI
# # =====================================

# Tk().withdraw()

# folder_path = askdirectory(
#     title="Select Amazon Invoice Folder"
# )

# if not folder_path:
#     print("No folder selected")
#     exit()

# # =====================================
# # RESULTS
# # =====================================

# results = []

# # =====================================
# # CLEAN TEXT
# # =====================================

# def clean_text(text):

#     if text is None:
#         return ""

#     text = text.replace("\n", " ")
#     text = re.sub(r'\s+', ' ', text)

#     return text.strip()

# # =====================================
# # CLEAN NUMBER
# # =====================================

# def clean_number(value):

#     if not value:
#         return 0.0

#     value = str(value)

#     value = value.replace("INR", "")
#     value = value.replace("-INR", "-")
#     value = value.replace(",", "")
#     value = value.strip()

#     try:
#         return float(value)

#     except:
#         return 0.0

# # =====================================
# # EXTRACT FUNCTION
# # =====================================

# def extract(pattern, text):

#     match = re.search(
#         pattern,
#         text,
#         re.IGNORECASE | re.DOTALL
#     )

#     if match:
#         return clean_text(match.group(1))

#     return ""

# # =====================================
# # PROCESS FILES
# # =====================================

# for root, dirs, files in os.walk(folder_path):

#     for file in files:

#         if not file.lower().endswith(".pdf"):
#             continue

#         full_path = os.path.join(root, file)

#         print(f"Processing: {file}")

#         try:

#             # =================================
#             # READ PDF
#             # =================================

#             doc = fitz.open(full_path)

#             full_text = ""

#             for page in doc:
#                 full_text += page.get_text()

#             normalized_text = clean_text(full_text)

#             # =================================
#             # REMOVE SECOND TABLE
#             # =================================

#             normalized_text = normalized_text.split(
#                 "Details of Fees to the above"
#             )[0]

#             # =================================
#             # DOCUMENT TYPE
#             # =================================

#             document_type = "Invoice"

#             if "Credit Note" in normalized_text:
#                 document_type = "Credit Note"

#             # =================================
#             # COMPANY DETAILS
#             # =================================

#             company_name = extract(

#                 r'(Amazon Seller Services Private Limited|Amazon Retail India Private Limited)',

#                 normalized_text
#             )

#             company_gstin = extract(

#                 r'GST Tax Registration No:\s*([A-Z0-9]+)',

#                 normalized_text
#             )

#             # =================================
#             # INVOICE DETAILS
#             # =================================

#             invoice_no = extract(

#                 r'(?:Invoice Number|Credit Note Number)\s*:\s*([A-Z0-9\-]+)',

#                 normalized_text
#             )

#             invoice_date = extract(

#                 r'(?:Invoice Date|Credit Note Date)\s*:\s*([0-9\/\-]+)',

#                 normalized_text
#             )

#             # =================================
#             # CUSTOMER DETAILS
#             # =================================

#             customer_name = extract(

#                 r'Bill to\s*Name:\s*(.*?)Address:',

#                 normalized_text
#             )

#             customer_name = clean_text(customer_name)

#             customer_gstin = extract(

#                 r'GSTIN:\s*([A-Z0-9]+)',

#                 normalized_text
#             )

#             # =================================
#             # INVOICE TOTAL
#             # =================================

#             invoice_total = extract(

#                 r'Total(?: Invoice amount)?\s*:?\s*INR\s*([0-9,]+\.\d+)',

#                 normalized_text
#             )

#             invoice_total = clean_number(invoice_total)

#             # =================================
#             # INVOICE ROWS
#             # =================================

#             if document_type == "Invoice":

#                 row_pattern = re.compile(

#                     r'(\d+)\.\s+'
#                     r'(\d{6})\s+'
#                     r'(.*?)\s+'
#                     r'INR\s*([0-9,]+\.\d+)'
#                     r'(.*?)'
#                     r'(?=\d+\.\s+\d{6}|\Z)',

#                     re.IGNORECASE | re.DOTALL
#                 )

#                 matches = row_pattern.findall(normalized_text)

#                 for row in matches:

#                     (
#                         serial_no,
#                         sac_code,
#                         description,
#                         fee_amount,
#                         tax_text
#                     ) = row

#                     description = clean_text(description)

#                     fee_amount = clean_number(fee_amount)

#                     # =================================
#                     # TAXES
#                     # =================================

#                     sgst_rate = 0.0
#                     sgst_amount = 0.0

#                     cgst_rate = 0.0
#                     cgst_amount = 0.0

#                     igst_rate = 0.0
#                     igst_amount = 0.0

#                     # SGST

#                     sgst_match = re.search(

#                         r'SGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',

#                         tax_text
#                     )

#                     if sgst_match:

#                         sgst_rate = clean_number(
#                             sgst_match.group(1)
#                         )

#                         sgst_amount = clean_number(
#                             sgst_match.group(2)
#                         )

#                     # CGST

#                     cgst_match = re.search(

#                         r'CGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',

#                         tax_text
#                     )

#                     if cgst_match:

#                         cgst_rate = clean_number(
#                             cgst_match.group(1)
#                         )

#                         cgst_amount = clean_number(
#                             cgst_match.group(2)
#                         )

#                     # IGST

#                     igst_match = re.search(

#                         r'IGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',

#                         tax_text
#                     )

#                     if igst_match:

#                         igst_rate = clean_number(
#                             igst_match.group(1)
#                         )

#                         igst_amount = clean_number(
#                             igst_match.group(2)
#                         )

#                     # =================================
#                     # LINE TOTAL
#                     # =================================

#                     line_total = round(

#                         fee_amount +
#                         sgst_amount +
#                         cgst_amount +
#                         igst_amount,

#                         2
#                     )

#                     # =================================
#                     # APPEND RESULTS
#                     # =================================

#                     results.append({

#                         "File Name": file,

#                         "Document Type": document_type,

#                         "Invoice No": invoice_no,
#                         "Invoice Date": invoice_date,

#                         "Company Name": company_name,
#                         "Company GSTIN": company_gstin,

#                         "Customer Name": customer_name,
#                         "Customer GSTIN": customer_gstin,

#                         "SAC Code": int(sac_code),

#                         "Description": description,

#                         "Fee Amount": fee_amount,

#                         "SGST Rate": sgst_rate,
#                         "SGST Amount": sgst_amount,

#                         "CGST Rate": cgst_rate,
#                         "CGST Amount": cgst_amount,

#                         "IGST Rate": igst_rate,
#                         "IGST Amount": igst_amount,

#                         "Line Total": line_total,

#                         "Invoice Total": invoice_total
#                     })

#             # =================================
#             # CREDIT NOTE ROWS
#             # =================================

#             else:

#                 credit_pattern = re.compile(

#                     r'(\d+)\.\s+'
#                     r'([A-Z0-9\-]+)\s+'
#                     r'([0-9\-]+)\s+'
#                     r'(\d{6})\s+'
#                     r'(.*?)'
#                     r'-INR\s*([0-9,]+\.\d+)'
#                     r'(.*?)'
#                     r'(?=\d+\.\s+[A-Z0-9\-]+|\Z)',

#                     re.IGNORECASE | re.DOTALL
#                 )

#                 matches = credit_pattern.findall(normalized_text)

#                 for row in matches:

#                     (
#                         serial_no,
#                         original_invoice,
#                         original_date,
#                         sac_code,
#                         description,
#                         fee_amount,
#                         tax_text
#                     ) = row

#                     description = clean_text(description)

#                     fee_amount = -abs(
#                         clean_number(fee_amount)
#                     )

#                     sgst_rate = 0.0
#                     sgst_amount = 0.0

#                     cgst_rate = 0.0
#                     cgst_amount = 0.0

#                     # SGST

#                     sgst_match = re.search(

#                         r'SGST\s*([0-9\.]+)%\s*-INR\s*([0-9,]+\.\d+)',

#                         tax_text
#                     )

#                     if sgst_match:

#                         sgst_rate = clean_number(
#                             sgst_match.group(1)
#                         )

#                         sgst_amount = -abs(
#                             clean_number(
#                                 sgst_match.group(2)
#                             )
#                         )

#                     # CGST

#                     cgst_match = re.search(

#                         r'CGST\s*([0-9\.]+)%\s*-INR\s*([0-9,]+\.\d+)',

#                         tax_text
#                     )

#                     if cgst_match:

#                         cgst_rate = clean_number(
#                             cgst_match.group(1)
#                         )

#                         cgst_amount = -abs(
#                             clean_number(
#                                 cgst_match.group(2)
#                             )
#                         )

#                     # =================================
#                     # LINE TOTAL
#                     # =================================

#                     line_total = round(

#                         fee_amount +
#                         sgst_amount +
#                         cgst_amount,

#                         2
#                     )

#                     # =================================
#                     # APPEND RESULTS
#                     # =================================

#                     results.append({

#                         "File Name": file,

#                         "Document Type": document_type,

#                         "Invoice No": invoice_no,
#                         "Invoice Date": invoice_date,

#                         "Original Invoice": original_invoice,
#                         "Original Invoice Date": original_date,

#                         "Company Name": company_name,
#                         "Company GSTIN": company_gstin,

#                         "Customer Name": customer_name,
#                         "Customer GSTIN": customer_gstin,

#                         "SAC Code": int(sac_code),

#                         "Description": description,

#                         "Fee Amount": fee_amount,

#                         "SGST Rate": sgst_rate,
#                         "SGST Amount": sgst_amount,

#                         "CGST Rate": cgst_rate,
#                         "CGST Amount": cgst_amount,

#                         "IGST Rate": 0.0,
#                         "IGST Amount": 0.0,

#                         "Line Total": line_total,

#                         "Invoice Total": -abs(invoice_total)
#                     })

#         except Exception as e:

#             print(f"Error processing {file}")
#             print(str(e))

# # =====================================
# # CREATE DATAFRAME
# # =====================================

# df = pd.DataFrame(results)

# # =====================================
# # OUTPUT FILE
# # =====================================

# output_file = os.path.join(
#     folder_path,
#     "Amazon_Invoice_Output.xlsx"
# )

# # =====================================
# # EXPORT EXCEL
# # =====================================

# with pd.ExcelWriter(
#     output_file,
#     engine="openpyxl"
# ) as writer:

#     df.to_excel(writer, index=False)

#     worksheet = writer.sheets["Sheet1"]

#     # =================================
#     # NUMBER FORMAT
#     # =================================

#     for row in worksheet.iter_rows(min_row=2):

#         for cell in row:

#             if isinstance(cell.value, (int, float)):

#                 cell.number_format = '0.00'

# # =====================================
# # SUCCESS POPUP
# # =====================================

# root = Tk()
# root.withdraw()

# messagebox.showinfo(
#     "Completed",
#     f"Amazon Invoice Extraction Completed!\n\nOutput File:\n{output_file}"
# )

# print("===================================")
# print("Amazon Extraction Completed")
# print("===================================")
# print(f"Output File: {output_file}")


# import fitz
# import pandas as pd
# import re
# import os
# import shutil

# from fastapi import FastAPI, UploadFile, File
# from fastapi.middleware.cors import CORSMiddleware
# from fastapi.responses import StreamingResponse
# from typing import List
# from io import BytesIO

# app = FastAPI()

# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"],
# )

# UPLOAD_DIR = "uploads"
# OUTPUT_DIR = "outputs"

# os.makedirs(UPLOAD_DIR, exist_ok=True)
# os.makedirs(OUTPUT_DIR, exist_ok=True)

# # =====================================
# # CLEAN TEXT
# # =====================================

# def clean_text(text):

#     if text is None:
#         return ""

#     text = text.replace("\n", " ")
#     text = re.sub(r"\s+", " ", text)

#     return text.strip()

# # =====================================
# # CLEAN NUMBER
# # =====================================

# def clean_number(value):

#     if not value:
#         return 0.0

#     value = str(value)

#     value = value.replace("INR", "")
#     value = value.replace("-INR", "-")
#     value = value.replace(",", "")
#     value = value.strip()

#     try:
#         return float(value)

#     except:
#         return 0.0

# # =====================================
# # EXTRACT FUNCTION
# # =====================================

# def extract(pattern, text):

#     match = re.search(
#         pattern,
#         text,
#         re.IGNORECASE | re.DOTALL
#     )

#     if match:
#         return clean_text(match.group(1))

#     return ""

# # =====================================
# # PROCESS PDF
# # =====================================

# def process_pdf(pdf_path):

#     results = []

#     file_name = os.path.basename(pdf_path)

#     doc = fitz.open(pdf_path)

#     full_text = ""

#     for page in doc:
#         full_text += page.get_text()

#     normalized_text = clean_text(full_text)

#     normalized_text = normalized_text.split(
#         "Details of Fees to the above"
#     )[0]

#     document_type = "Invoice"

#     if "Credit Note" in normalized_text:
#         document_type = "Credit Note"

#     company_name = extract(
#         r'(Amazon Seller Services Private Limited|Amazon Retail India Private Limited)',
#         normalized_text
#     )

#     company_gstin = extract(
#         r'GST Tax Registration No:\s*([A-Z0-9]+)',
#         normalized_text
#     )

#     invoice_no = extract(
#         r'(?:Invoice Number|Credit Note Number)\s*:\s*([A-Z0-9\-]+)',
#         normalized_text
#     )

#     invoice_date = extract(
#         r'(?:Invoice Date|Credit Note Date)\s*:\s*([0-9\/\-]+)',
#         normalized_text
#     )

#     customer_name = extract(
#         r'Bill to\s*Name:\s*(.*?)Address:',
#         normalized_text
#     )

#     customer_gstin = extract(
#         r'GSTIN:\s*([A-Z0-9]+)',
#         normalized_text
#     )

#     invoice_total = extract(
#         r'Total(?: Invoice amount)?\s*:?\s*INR\s*([0-9,]+\.\d+)',
#         normalized_text
#     )

#     invoice_total = clean_number(invoice_total)

#     row_pattern = re.compile(
#         r'(\d+)\.\s+'
#         r'(\d{6})\s+'
#         r'(.*?)\s+'
#         r'INR\s*([0-9,]+\.\d+)'
#         r'(.*?)'
#         r'(?=\d+\.\s+\d{6}|\Z)',
#         re.IGNORECASE | re.DOTALL
#     )

#     matches = row_pattern.findall(normalized_text)

#     for row in matches:

#         (
#             serial_no,
#             sac_code,
#             description,
#             fee_amount,
#             tax_text
#         ) = row

#         description = clean_text(description)

#         fee_amount = clean_number(fee_amount)

#         sgst_rate = 0.0
#         sgst_amount = 0.0

#         cgst_rate = 0.0
#         cgst_amount = 0.0

#         igst_rate = 0.0
#         igst_amount = 0.0

#         sgst_match = re.search(
#             r'SGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
#             tax_text
#         )

#         if sgst_match:

#             sgst_rate = clean_number(
#                 sgst_match.group(1)
#             )

#             sgst_amount = clean_number(
#                 sgst_match.group(2)
#             )

#         cgst_match = re.search(
#             r'CGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
#             tax_text
#         )

#         if cgst_match:

#             cgst_rate = clean_number(
#                 cgst_match.group(1)
#             )

#             cgst_amount = clean_number(
#                 cgst_match.group(2)
#             )

#         line_total = round(
#             fee_amount +
#             sgst_amount +
#             cgst_amount,
#             2
#         )

#         results.append({

#             "File Name": file_name,
#             "Document Type": document_type,

#             "Invoice No": invoice_no,
#             "Invoice Date": invoice_date,

#             "Company Name": company_name,
#             "Company GSTIN": company_gstin,

#             "Customer Name": customer_name,
#             "Customer GSTIN": customer_gstin,

#             "SAC Code": sac_code,

#             "Description": description,

#             "Fee Amount": fee_amount,

#             "SGST Rate": sgst_rate,
#             "SGST Amount": sgst_amount,

#             "CGST Rate": cgst_rate,
#             "CGST Amount": cgst_amount,

#             "Line Total": line_total,

#             "Invoice Total": invoice_total
#         })

#     return results

# # =====================================
# # API
# # =====================================

# @app.post("/upload")

# async def upload_files(
#     files: List[UploadFile] = File(...)
# ):

#     all_results = []

#     for file in files:

#         if not file.filename.lower().endswith(".pdf"):
#             continue

#         save_path = os.path.join(
#             UPLOAD_DIR,
#             file.filename
#         )

#         with open(save_path, "wb") as buffer:
#             shutil.copyfileobj(file.file, buffer)

#         extracted = process_pdf(save_path)

#         all_results.extend(extracted)

#     df = pd.DataFrame(all_results)

#     excel_buffer = BytesIO()

#     with pd.ExcelWriter(
#         excel_buffer,
#         engine="openpyxl"
#     ) as writer:

#         df.to_excel(
#             writer,
#             index=False,
#             sheet_name="Invoices"
#         )

#         worksheet = writer.sheets["Invoices"]

#         for column in worksheet.columns:

#             max_length = 0

#             column_letter = column[0].column_letter

#             for cell in column:

#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass

#             worksheet.column_dimensions[
#                 column_letter
#             ].width = max_length + 5

#     excel_buffer.seek(0)

#     return StreamingResponse(
#         excel_buffer,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Content-Disposition":
#             "attachment; filename=Amazon_Invoice_Output.xlsx"
#         }
#     )

# from fastapi import FastAPI, UploadFile, File
# from fastapi.middleware.cors import CORSMiddleware

# from typing import List

# import os
# import shutil
# import fitz
# import pandas as pd
# import re

# app = FastAPI()

# # =========================================
# # CORS
# # =========================================

# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"],
# )

# # =========================================
# # CREATE FOLDERS
# # =========================================

# UPLOAD_FOLDER = "uploads"
# OUTPUT_FOLDER = "output"

# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# # =========================================
# # CLEAN TEXT
# # =========================================

# def clean_text(text):

#     if text is None:
#         return ""

#     text = text.replace("\n", " ")
#     text = re.sub(r"\s+", " ", text)

#     return text.strip()

# # =========================================
# # CLEAN NUMBER
# # =========================================

# def clean_number(value):

#     if not value:
#         return 0.0

#     value = str(value)

#     value = value.replace("INR", "")
#     value = value.replace("-INR", "-")
#     value = value.replace(",", "")
#     value = value.strip()

#     try:
#         return float(value)

#     except:
#         return 0.0

# # =========================================
# # EXTRACT
# # =========================================

# def extract(pattern, text):

#     match = re.search(
#         pattern,
#         text,
#         re.IGNORECASE | re.DOTALL
#     )

#     if match:
#         return clean_text(match.group(1))

#     return ""

# # =========================================
# # UPLOAD API
# # =========================================

# @app.post("/upload")
# async def upload_files(
#     files: List[UploadFile] = File(...)
# ):

#     results = []

#     # =====================================
#     # SAVE FILES
#     # =====================================

#     for file in files:

#         try:

#             # =================================
#             # SAFE FILE NAME
#             # =================================

#             safe_filename = os.path.basename(
#                 file.filename
#             )

#             safe_filename = safe_filename.replace("/", "_")
#             safe_filename = safe_filename.replace("\\", "_")

#             save_path = os.path.join(
#                 UPLOAD_FOLDER,
#                 safe_filename
#             )

#             # =================================
#             # SAVE PDF
#             # =================================

#             with open(save_path, "wb") as buffer:
#                 shutil.copyfileobj(
#                     file.file,
#                     buffer
#                 )

#             print(f"Saved: {safe_filename}")

#             # =================================
#             # READ PDF
#             # =================================

#             doc = fitz.open(save_path)

#             full_text = ""

#             for page in doc:
#                 full_text += page.get_text()

#             normalized_text = clean_text(
#                 full_text
#             )

#             normalized_text = normalized_text.split(
#                 "Details of Fees to the above"
#             )[0]

#             # =================================
#             # DOCUMENT TYPE
#             # =================================

#             document_type = "Invoice"

#             if "Credit Note" in normalized_text:
#                 document_type = "Credit Note"

#             # =================================
#             # COMPANY
#             # =================================

#             company_name = extract(
#                 r'(Amazon Seller Services Private Limited|Amazon Retail India Private Limited)',
#                 normalized_text
#             )

#             company_gstin = extract(
#                 r'GST Tax Registration No:\s*([A-Z0-9]+)',
#                 normalized_text
#             )

#             # =================================
#             # INVOICE
#             # =================================

#             invoice_no = extract(
#                 r'(?:Invoice Number|Credit Note Number)\s*:\s*([A-Z0-9\-]+)',
#                 normalized_text
#             )

#             invoice_date = extract(
#                 r'(?:Invoice Date|Credit Note Date)\s*:\s*([0-9\/\-]+)',
#                 normalized_text
#             )

#             invoice_total = extract(
#                 r'Total(?: Invoice amount)?\s*:?\s*INR\s*([0-9,]+\.\d+)',
#                 normalized_text
#             )

#             invoice_total = clean_number(
#                 invoice_total
#             )

#             # =================================
#             # CUSTOMER
#             # =================================

#             customer_name = extract(
#                 r'Bill to\s*Name:\s*(.*?)Address:',
#                 normalized_text
#             )

#             customer_gstin = extract(
#                 r'GSTIN:\s*([A-Z0-9]+)',
#                 normalized_text
#             )

#             # =================================
#             # ROWS
#             # =================================

#             row_pattern = re.compile(

#                 r'(\d+)\.\s+'
#                 r'(\d{6})\s+'
#                 r'(.*?)\s+'
#                 r'INR\s*([0-9,]+\.\d+)'
#                 r'(.*?)'
#                 r'(?=\d+\.\s+\d{6}|\Z)',

#                 re.IGNORECASE | re.DOTALL
#             )

#             matches = row_pattern.findall(
#                 normalized_text
#             )

#             for row in matches:

#                 (
#                     serial_no,
#                     sac_code,
#                     description,
#                     fee_amount,
#                     tax_text
#                 ) = row

#                 description = clean_text(
#                     description
#                 )

#                 fee_amount = clean_number(
#                     fee_amount
#                 )

#                 sgst_rate = 0.0
#                 sgst_amount = 0.0

#                 cgst_rate = 0.0
#                 cgst_amount = 0.0

#                 igst_rate = 0.0
#                 igst_amount = 0.0

#                 # =============================
#                 # SGST
#                 # =============================

#                 sgst_match = re.search(
#                     r'SGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
#                     tax_text
#                 )

#                 if sgst_match:

#                     sgst_rate = clean_number(
#                         sgst_match.group(1)
#                     )

#                     sgst_amount = clean_number(
#                         sgst_match.group(2)
#                     )

#                 # =============================
#                 # CGST
#                 # =============================

#                 cgst_match = re.search(
#                     r'CGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
#                     tax_text
#                 )

#                 if cgst_match:

#                     cgst_rate = clean_number(
#                         cgst_match.group(1)
#                     )

#                     cgst_amount = clean_number(
#                         cgst_match.group(2)
#                     )

#                 # =============================
#                 # IGST
#                 # =============================

#                 igst_match = re.search(
#                     r'IGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
#                     tax_text
#                 )

#                 if igst_match:

#                     igst_rate = clean_number(
#                         igst_match.group(1)
#                     )

#                     igst_amount = clean_number(
#                         igst_match.group(2)
#                     )

#                 # =============================
#                 # LINE TOTAL
#                 # =============================

#                 line_total = round(

#                     fee_amount +
#                     sgst_amount +
#                     cgst_amount +
#                     igst_amount,

#                     2
#                 )

#                 # =============================
#                 # APPEND
#                 # =============================

#                 results.append({

#                     "File Name": safe_filename,

#                     "Document Type": document_type,

#                     "Invoice No": invoice_no,
#                     "Invoice Date": invoice_date,

#                     "Company Name": company_name,
#                     "Company GSTIN": company_gstin,

#                     "Customer Name": customer_name,
#                     "Customer GSTIN": customer_gstin,

#                     "SAC Code": sac_code,

#                     "Description": description,

#                     "Fee Amount": fee_amount,

#                     "SGST Rate": sgst_rate,
#                     "SGST Amount": sgst_amount,

#                     "CGST Rate": cgst_rate,
#                     "CGST Amount": cgst_amount,

#                     "IGST Rate": igst_rate,
#                     "IGST Amount": igst_amount,

#                     "Line Total": line_total,

#                     "Invoice Total": invoice_total
#                 })

#         except Exception as e:

#             print(f"ERROR: {file.filename}")
#             print(str(e))

#     # =====================================
#     # CREATE DATAFRAME
#     # =====================================

#     df = pd.DataFrame(results)

#     # =====================================
#     # OUTPUT EXCEL
#     # =====================================

#     output_file = os.path.join(
#         OUTPUT_FOLDER,
#         "Amazon_Invoice_Output.xlsx"
#     )

#     # with pd.ExcelWriter(
#     #     output_file,
#     #     engine="openpyxl"
#     # ) as writer:

#     #     df.to_excel(
#     #         writer,
#     #         index=False
#     #     )
    
#     if df.empty:

#         df = pd.DataFrame([
#         {
#             "Message": "No Data Found"
#         }
#     ])

#     # save excel properly
#     df.to_excel(
#         output_file,
#         index=False,
#         engine="openpyxl"
#     )

#     # reopen workbook to validate
#     wb = load_workbook(output_file)
#     wb.save(output_file)
#     wb.close()

#     # =====================================
#     # RESPONSE
#     # =====================================

#     return {
#         "message": "Success",
#         "excel_file": output_file,
#         "total_rows": len(df)
#     }
    

from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

from typing import List

import os
import shutil
import fitz
import pandas as pd
import re

from openpyxl import load_workbook

app = FastAPI()

# =====================================================
# CORS
# =====================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =====================================================
# FOLDERS
# =====================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# =====================================================
# CLEAN TEXT
# =====================================================

def clean_text(text):

    if text is None:
        return ""

    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)

    return text.strip()

# =====================================================
# CLEAN NUMBER
# =====================================================

def clean_number(value):

    if not value:
        return 0.0

    value = str(value)

    value = value.replace("INR", "")
    value = value.replace("-INR", "-")
    value = value.replace(",", "")
    value = value.strip()

    try:
        return float(value)

    except:
        return 0.0

# =====================================================
# EXTRACT REGEX
# =====================================================

def extract(pattern, text):

    match = re.search(
        pattern,
        text,
        re.IGNORECASE | re.DOTALL
    )

    if match:
        return clean_text(match.group(1))

    return ""

# =====================================================
# HOME
# =====================================================

@app.get("/")
def home():

    return {
        "message": "Amazon Invoice OCR API Running"
    }

# =====================================================
# UPLOAD API
# =====================================================

@app.post("/upload")
async def upload_files(
    files: List[UploadFile] = File(...)
):

    results = []

    # =================================================
    # LOOP FILES
    # =================================================

    for file in files:

        try:

            print(f"Processing: {file.filename}")

            # =============================================
            # SAFE FILE NAME
            # =============================================

            safe_filename = os.path.basename(file.filename)

            safe_filename = safe_filename.replace("/", "_")
            safe_filename = safe_filename.replace("\\", "_")

            # =============================================
            # SAVE PATH
            # =============================================

            save_path = os.path.join(
                UPLOAD_FOLDER,
                safe_filename
            )

            # =============================================
            # SAVE FILE
            # =============================================

            with open(save_path, "wb") as buffer:

                shutil.copyfileobj(
                    file.file,
                    buffer
                )

            print(f"Saved Successfully: {save_path}")

            # =============================================
            # OPEN PDF
            # =============================================

            doc = fitz.open(save_path)

            full_text = ""

            for page in doc:

                full_text += page.get_text()

            doc.close()

            # =============================================
            # CLEAN TEXT
            # =============================================

            normalized_text = clean_text(full_text)

            normalized_text = normalized_text.split(
                "Details of Fees to the above"
            )[0]

            # =============================================
            # DOCUMENT TYPE
            # =============================================

            document_type = "Invoice"

            if "Credit Note" in normalized_text:
                document_type = "Credit Note"

            # =============================================
            # COMPANY
            # =============================================

            company_name = extract(
                r'(Amazon Seller Services Private Limited|Amazon Retail India Private Limited)',
                normalized_text
            )

            company_gstin = extract(
                r'GST Tax Registration No:\s*([A-Z0-9]+)',
                normalized_text
            )

            # =============================================
            # INVOICE DETAILS
            # =============================================

            invoice_no = extract(
                r'(?:Invoice Number|Credit Note Number)\s*:\s*([A-Z0-9\-]+)',
                normalized_text
            )

            invoice_date = extract(
                r'(?:Invoice Date|Credit Note Date)\s*:\s*([0-9\/\-]+)',
                normalized_text
            )

            invoice_total = extract(
                r'Total(?: Invoice amount)?\s*:?\s*INR\s*([0-9,]+\.\d+)',
                normalized_text
            )

            invoice_total = clean_number(
                invoice_total
            )

            # =============================================
            # CUSTOMER
            # =============================================

            customer_name = extract(
                r'Bill to\s*Name:\s*(.*?)Address:',
                normalized_text
            )

            customer_gstin = extract(
                r'GSTIN:\s*([A-Z0-9]+)',
                normalized_text
            )

            # =============================================
            # ITEM ROWS
            # =============================================

            row_pattern = re.compile(

                r'(\d+)\.\s+'
                r'(\d{6})\s+'
                r'(.*?)\s+'
                r'INR\s*([0-9,]+\.\d+)'
                r'(.*?)'
                r'(?=\d+\.\s+\d{6}|\Z)',

                re.IGNORECASE | re.DOTALL
            )

            matches = row_pattern.findall(
                normalized_text
            )

            print(f"Rows Found: {len(matches)}")

            # =============================================
            # LOOP ROWS
            # =============================================

            for row in matches:

                (
                    serial_no,
                    sac_code,
                    description,
                    fee_amount,
                    tax_text
                ) = row

                description = clean_text(
                    description
                )

                fee_amount = clean_number(
                    fee_amount
                )

                sgst_rate = 0.0
                sgst_amount = 0.0

                cgst_rate = 0.0
                cgst_amount = 0.0

                igst_rate = 0.0
                igst_amount = 0.0

                # =========================================
                # SGST
                # =========================================

                sgst_match = re.search(
                    r'SGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
                    tax_text
                )

                if sgst_match:

                    sgst_rate = clean_number(
                        sgst_match.group(1)
                    )

                    sgst_amount = clean_number(
                        sgst_match.group(2)
                    )

                # =========================================
                # CGST
                # =========================================

                cgst_match = re.search(
                    r'CGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
                    tax_text
                )

                if cgst_match:

                    cgst_rate = clean_number(
                        cgst_match.group(1)
                    )

                    cgst_amount = clean_number(
                        cgst_match.group(2)
                    )

                # =========================================
                # IGST
                # =========================================

                igst_match = re.search(
                    r'IGST\s*([0-9\.]+)%\s*INR\s*([0-9,]+\.\d+)',
                    tax_text
                )

                if igst_match:

                    igst_rate = clean_number(
                        igst_match.group(1)
                    )

                    igst_amount = clean_number(
                        igst_match.group(2)
                    )

                # =========================================
                # LINE TOTAL
                # =========================================

                line_total = round(

                    fee_amount +
                    sgst_amount +
                    cgst_amount +
                    igst_amount,

                    2
                )

                # =========================================
                # APPEND RESULT
                # =========================================

                results.append({

                    "File Name": safe_filename,

                    "Document Type": document_type,

                    "Invoice No": invoice_no,
                    "Invoice Date": invoice_date,

                    "Company Name": company_name,
                    "Company GSTIN": company_gstin,

                    "Customer Name": customer_name,
                    "Customer GSTIN": customer_gstin,

                    "SAC Code": sac_code,

                    "Description": description,

                    "Fee Amount": fee_amount,

                    "SGST Rate": sgst_rate,
                    "SGST Amount": sgst_amount,

                    "CGST Rate": cgst_rate,
                    "CGST Amount": cgst_amount,

                    "IGST Rate": igst_rate,
                    "IGST Amount": igst_amount,

                    "Line Total": line_total,

                    "Invoice Total": invoice_total
                })

        except Exception as e:

            print("=" * 50)
            print(f"ERROR FILE: {file.filename}")
            print(str(e))
            print("=" * 50)

    # =================================================
    # CREATE DATAFRAME
    # =================================================

    df = pd.DataFrame(results)

    # =================================================
    # EMPTY CHECK
    # =================================================

    if df.empty:

        df = pd.DataFrame([
            {
                "Message": "No Data Found"
            }
        ])

    # =================================================
    # OUTPUT EXCEL FILE
    # =================================================

    output_file = os.path.join(
        OUTPUT_FOLDER,
        "Amazon_Invoice_Output.xlsx"
    )

    # DELETE OLD FILE
    if os.path.exists(output_file):
        os.remove(output_file)

    # =================================================
    # SAVE EXCEL
    # =================================================

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="Invoices"
        )

    # =================================================
    # VALIDATE EXCEL
    # =================================================

    wb = load_workbook(output_file)

    ws = wb.active

    # AUTO WIDTH
    for column_cells in ws.columns:

        length = max(

            len(str(cell.value))
            if cell.value else 0

            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 5

    wb.save(output_file)
    wb.close()

    print(f"Excel Created: {output_file}")

    # =================================================
    # RETURN FILE
    # =================================================

    return FileResponse(
        path=output_file,
        filename="Amazon_Invoice_Output.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    