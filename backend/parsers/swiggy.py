# from database.models import User 
# from fastapi import Depends

# from sqlalchemy.orm import Session

# from auth.dependencies import get_current_user

# from database.database import get_db

# from database.models import UploadAnalytics


# from fastapi import FastAPI, UploadFile, File
# from fastapi.middleware.cors import CORSMiddleware
# from fastapi.responses import FileResponse
# from typing import List
# import fitz
# import pandas as pd
# import re
# import os
# import shutil
# from openpyxl import load_workbook

# # app = FastAPI()

# # app.add_middleware(
# #     CORSMiddleware,
# #     allow_origins=["*"],
# #     allow_credentials=True,
# #     allow_methods=["*"],
# #     allow_headers=["*"],
# # )

# # BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# BASE_DIR = os.path.dirname(
#     os.path.dirname(os.path.abspath(__file__))
# )

# UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
# OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# VALID_DESCRIPTIONS = [
#     "Service Fee",
#     "Marketing services",
#     "Lead Gen Fees",
#     "Recovery for additional cashbacks given to customers by restaurants",
#     "Packing Charges",
#     "Delivery Charges",
#     "Advertisement Fees",
#     "Cancellation Charges",
#     "Convenience Fee"
# ]

# def clean_text(text):
#     if text is None:
#         return ""
#     text = str(text).replace("\n", " ")
#     text = re.sub(r"\s+", " ", text)
#     return text.strip()

# def clean_number(value):
#     if value is None:
#         return 0.0
#     value = str(value).strip()
#     value = value.replace(",", "").replace("INR", "").replace("Rs.", "")
#     try:
#         return float(value)
#     except:
#         return 0.0

# def extract(pattern, text):
#     match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
#     return clean_text(match.group(1)) if match else ""

# def process_swiggy_pdf(full_path, file_name):
#     results = []

#     doc = fitz.open(full_path)
#     full_text = ""

#     for page in doc:
#         full_text += page.get_text()

#     doc.close()

#     text = clean_text(full_text)

#     invoice_no = extract(r'Invoice Number\s*:\s*([A-Z0-9\-]+)', text)
#     invoice_date = extract(r'Invoice Date\s*:\s*([0-9:\-\s]+)', text)

#     service_period_full = extract(
#         r'Service Period\s*:\s*(.*?)PO Number',
#         text
#     )

#     service_period_parts = re.split(
#         r'Address\s*:',
#         service_period_full,
#         flags=re.IGNORECASE
#     )

#     service_period = service_period_parts[0].strip()
#     service_address = ""

#     if len(service_period_parts) > 1:
#         service_address = service_period_parts[1].strip()

#     irn = extract(r'IRN\s*:\s*([A-Za-z0-9]+)', text)

#     company_name = "Swiggy Limited"

#     company_match = re.search(
#         r'GSTIN\s*:\s*([A-Z0-9]+)',
#         text,
#         re.IGNORECASE
#     )

#     company_gstin = company_match.group(1).strip() if company_match else ""

#     customer_name = extract(
#         r'Legal Name\s*:\s*(.*?)Restaurant',
#         text
#     )

#     restaurant_name = extract(
#         r'Restaurant\s*/\s*Store Name\s*:\s*(.*?)State Code',
#         text
#     )

#     restaurant_id = extract(
#         r'Restaurant\s*/\s*Store ID\s*:\s*([0-9]+)',
#         text
#     )

#     gstin_matches = re.findall(
#         r'GSTIN\s*:\s*([A-Z0-9]+)',
#         text,
#         re.IGNORECASE
#     )

#     customer_gstin = gstin_matches[1] if len(gstin_matches) >= 2 else ""

#     grand_total = clean_number(
#         extract(r'Grand Total\s*([0-9,\.]+)', text)
#     )

#     reimbursement_discount = clean_number(
#         extract(
#             r'Other Charges\s*-\s*Reimbursement\s*of\s*Discount\s*([0-9,\.]+)',
#             text
#         )
#     )

#     table_start = text.find("Sr.")

#     if table_start == -1:
#         return []

#     table_text = text[table_start:]

#     line_pattern = re.compile(
#         r'(\d+)\s+'
#         r'([A-Za-z0-9 \-\&\/\(\)\.\,\%]+?)\s+'
#         r'(\d{6})\s+'
#         r'OTH\s+'
#         r'(\d+)\s+'
#         r'([0-9,\.\-]+)\s+'
#         r'([0-9,\.\-]+)\s+'
#         r'([0-9,\.\-]+)\s+'
#         r'([0-9,\.\-]+)\s+'
#         r'([0-9\.]*)\s+'
#         r'([0-9,\.\-]*)\s+'
#         r'([0-9\.]*)\s+'
#         r'([0-9,\.\-]*)\s+'
#         r'([0-9\.]*)\s+'
#         r'([0-9,\.\-]*)',
#         re.IGNORECASE
#     )

#     matches = line_pattern.findall(table_text)

#     if not matches:
#         return []

#     invoice_row = {
#         "File Name": file_name,
#         "Invoice No": invoice_no,
#         "Invoice Date": invoice_date,
#         "Service Period": service_period,
#         "Service Address": service_address,
#         "IRN": irn,
#         "Company Name": company_name,
#         "Company GSTIN": company_gstin,
#         "Customer Name": customer_name,
#         "Restaurant Name": restaurant_name,
#         "Restaurant ID": restaurant_id,
#         "Customer GSTIN": customer_gstin,
#         "Taxable Amount": 0.0,
#         "CGST Amount": 0.0,
#         "SGST Amount": 0.0,
#         "IGST Amount": 0.0,
#         "Other Charges - Reimbursement of Discount": reimbursement_discount,
#         "Grand Total": grand_total
#     }

#     for row in matches:
#         description = clean_text(row[1])
#         assessable_value = clean_number(row[7])
#         cgst_amount = clean_number(row[9])
#         sgst_amount = clean_number(row[11])
#         igst_amount = clean_number(row[13])

#         final_description = ""

#         for desc in VALID_DESCRIPTIONS:
#             if desc.lower() in description.lower():
#                 final_description = desc
#                 break

#         if not final_description:
#             continue

#         invoice_row["Taxable Amount"] += assessable_value
#         invoice_row["CGST Amount"] += cgst_amount
#         invoice_row["SGST Amount"] += sgst_amount
#         invoice_row["IGST Amount"] += igst_amount

#         if final_description not in invoice_row:
#             invoice_row[final_description] = 0.0

#         invoice_row[final_description] += assessable_value

#     results.append(invoice_row)
#     return results


# async def upload_swiggy(files: List[UploadFile] = File(...),

#      db: Session = Depends(get_db),

#     current_user: User = Depends(
#         get_current_user
#     )
# ):
#     files_count = len(files)
    
#     results = []

#     for file in files:
#         safe_filename = os.path.basename(file.filename)
#         save_path = os.path.join(UPLOAD_FOLDER, safe_filename)

#         with open(save_path, "wb") as buffer:
#             shutil.copyfileobj(file.file, buffer)

#         try:
#             results.extend(
#                 process_swiggy_pdf(save_path, safe_filename)
#             )
#         except Exception as e:
#             print(e)

#     df = pd.DataFrame(results)

#     if df.empty:
#         df = pd.DataFrame([{"Message": "No Data Found"}])

#     output_file = os.path.join(
#         OUTPUT_FOLDER,
#         "Swiggy_Invoice_Output.xlsx"
#     )

#     with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False)

#     wb = load_workbook(output_file)
#     ws = wb.active

#     for col in ws.columns:
#         width = max(len(str(c.value)) if c.value else 0 for c in col)
#         ws.column_dimensions[col[0].column_letter].width = width + 5

#     wb.save(output_file)
#     wb.close()
    
#     if files_count > 0:

#         analytics = UploadAnalytics(

#             user_id=current_user.id,

#             source="swiggy",

#             pdf_count=files_count
#         )

#         db.add(analytics)

#         db.commit()

#     return FileResponse(
#         path=output_file,
#         filename="Swiggy_Invoice_Output.xlsx",
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )






import sys
import os
import fitz
import pandas as pd
import re
import shutil
from typing import List
from openpyxl import load_workbook

from fastapi import Depends, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from database.database import get_db
from database.models import UploadAnalytics

BASE_DIR = os.path.dirname(
    os.path.dirname(os.path.abspath(__file__))
)

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

VALID_DESCRIPTIONS = [
    "Service Fee",
    "Marketing services",
    "Lead Gen Fees",
    "Recovery for additional cashbacks given to customers by restaurants",
    "Packing Charges",
    "Delivery Charges",
    "Advertisement Fees",
    "Cancellation Charges",
    "Convenience Fee"
]

def clean_text(text):
    if text is None:
        return ""
    text = str(text).replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def clean_number(value):
    if value is None:
        return 0.0
    value = str(value).strip()
    value = value.replace(",", "").replace("INR", "").replace("Rs.", "")
    try:
        return float(value)
    except:
        return 0.0

def extract(pattern, text):
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return clean_text(match.group(1)) if match else ""

def process_swiggy_pdf(full_path, file_name):
    results = []

    doc = fitz.open(full_path)
    full_text = ""

    for page in doc:
        full_text += page.get_text()

    doc.close()

    text = clean_text(full_text)

    invoice_no = extract(r'Invoice Number\s*:\s*([A-Z0-9\-]+)', text)
    invoice_date = extract(r'Invoice Date\s*:\s*([0-9:\-\s]+)', text)

    service_period_full = extract(
        r'Service Period\s*:\s*(.*?)PO Number',
        text
    )

    service_period_parts = re.split(
        r'Address\s*:',
        service_period_full,
        flags=re.IGNORECASE
    )

    service_period = service_period_parts[0].strip()
    service_address = ""

    if len(service_period_parts) > 1:
        service_address = service_period_parts[1].strip()

    irn = extract(r'IRN\s*:\s*([A-Za-z0-9]+)', text)

    company_name = "Swiggy Limited"

    company_match = re.search(
        r'GSTIN\s*:\s*([A-Z0-9]+)',
        text,
        re.IGNORECASE
    )

    company_gstin = company_match.group(1).strip() if company_match else ""

    customer_name = extract(
        r'Legal Name\s*:\s*(.*?)Restaurant',
        text
    )

    restaurant_name = extract(
        r'Restaurant\s*/\s*Store Name\s*:\s*(.*?)State Code',
        text
    )

    restaurant_id = extract(
        r'Restaurant\s*/\s*Store ID\s*:\s*([0-9]+)',
        text
    )

    gstin_matches = re.findall(
        r'GSTIN\s*:\s*([A-Z0-9]+)',
        text,
        re.IGNORECASE
    )

    customer_gstin = gstin_matches[1] if len(gstin_matches) >= 2 else ""

    grand_total = clean_number(
        extract(r'Grand Total\s*([0-9,\.]+)', text)
    )

    reimbursement_discount = clean_number(
        extract(
            r'Other Charges\s*-\s*Reimbursement\s*of\s*Discount\s*([0-9,\.]+)',
            text
        )
    )

    table_start = text.find("Sr.")

    if table_start == -1:
        return []

    table_text = text[table_start:]

    line_pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Za-z0-9 \-\&\/\(\)\.\,\%]+?)\s+'
        r'(\d{6})\s+'
        r'OTH\s+'
        r'(\d+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9,\.\-]+)\s+'
        r'([0-9\.]*)\s+'
        r'([0-9,\.\-]*)\s+'
        r'([0-9\.]*)\s+'
        r'([0-9,\.\-]*)\s+'
        r'([0-9\.]*)\s+'
        r'([0-9,\.\-]*)',
        re.IGNORECASE
    )

    matches = line_pattern.findall(table_text)

    if not matches:
        return []

    invoice_row = {
        "File Name": file_name,
        "Invoice No": invoice_no,
        "Invoice Date": invoice_date,
        "Service Period": service_period,
        "Service Address": service_address,
        "IRN": irn,
        "Company Name": company_name,
        "Company GSTIN": company_gstin,
        "Customer Name": customer_name,
        "Restaurant Name": restaurant_name,
        "Restaurant ID": restaurant_id,
        "Customer GSTIN": customer_gstin,
        "Taxable Amount": 0.0,
        "CGST Amount": 0.0,
        "SGST Amount": 0.0,
        "IGST Amount": 0.0,
        "Other Charges - Reimbursement of Discount": reimbursement_discount,
        "Grand Total": grand_total
    }

    for row in matches:
        description = clean_text(row[1])
        assessable_value = clean_number(row[7])
        cgst_amount = clean_number(row[9])
        sgst_amount = clean_number(row[11])
        igst_amount = clean_number(row[13])

        final_description = ""

        for desc in VALID_DESCRIPTIONS:
            if desc.lower() in description.lower():
                final_description = desc
                break

        if not final_description:
            continue

        invoice_row["Taxable Amount"] += assessable_value
        invoice_row["CGST Amount"] += cgst_amount
        invoice_row["SGST Amount"] += sgst_amount
        invoice_row["IGST Amount"] += igst_amount

        if final_description not in invoice_row:
            invoice_row[final_description] = 0.0

        invoice_row[final_description] += assessable_value

    results.append(invoice_row)
    return results


async def upload_swiggy(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user = None  # Removed Pydantic/SQLAlchemy model hint to prevent Uvicorn startup crash
):
    files_count = len(files)
    results = []

    for file in files:
        safe_filename = os.path.basename(file.filename)
        save_path = os.path.join(UPLOAD_FOLDER, safe_filename)

        with open(save_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        try:
            results.extend(
                process_swiggy_pdf(save_path, safe_filename)
            )
        except Exception as e:
            print(f"Error processing file {safe_filename}: {str(e)}")

    df = pd.DataFrame(results)

    if df.empty:
        df = pd.DataFrame([{"Message": "No Data Found"}])

    output_file = os.path.join(
        OUTPUT_FOLDER,
        "Swiggy_Invoice_Output.xlsx"
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = width + 5

    wb.save(output_file)
    wb.close()
    
    # Save analytics in Neon DB if user is present
    if files_count > 0 and current_user and hasattr(current_user, 'id'):
        analytics = UploadAnalytics(
            user_id=current_user.id,
            source="swiggy",
            pdf_count=files_count
        )
        db.add(analytics)
        db.commit()

    return FileResponse(
        path=output_file,
        filename="Swiggy_Invoice_Output.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )