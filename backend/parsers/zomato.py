    
# import os
# import re
# import shutil
# from datetime import datetime
# from tempfile import NamedTemporaryFile
# from typing import List

# import pandas as pd
# import pdfplumber
# from fastapi import Depends, FastAPI, File, UploadFile
# from fastapi.middleware.cors import CORSMiddleware
# from fastapi.responses import FileResponse
# from openpyxl import load_workbook
# from sqlalchemy.orm import Session

# # Database and Auth modules as per your system structure
# from auth.dependencies import get_current_user
# from database.database import get_db
# from database.models import UploadAnalytics

# # Base directories only for saving incoming uploads
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# # =====================================================
# # HELPER FUNCTIONS
# # =====================================================
# def extract(pattern, text):
#     match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
#     return match.group(1).strip() if match else ""


# def to_float(value):
#     if not value:
#         return 0.0
#     try:
#         return float(value.replace(",", "").strip())
#     except:
#         return 0.0


# # =====================================================
# # CORE PDF PARSER FUNCTION
# # =====================================================
# def process_pdf(pdf_file):
#     try:
#         with pdfplumber.open(pdf_file) as pdf:
#             text = pdf.pages[0].extract_text() or ""

#         invoice_no = extract(r"Invoice No:\s*([A-Z0-9\-]+)", text)
#         invoice_date = extract(r"Invoice Date:\s*([0-9\-]+)", text)
#         invoice_period = extract(r"Period:\s*(.*?)\n", text)
#         merchant_name = extract(r"Entity Name:\s*(.*?)\s*State:", text)
#         merchant_id = extract(r"Merchant ID:\s*(\d+)", text)
#         restaurant_id = extract(r"Restaurant ID:\s*(\d+)", text)

#         gstin_matches = re.findall(r"GSTIN:\s*([0-9A-Z]{15})", text)
#         company_gstin = gstin_matches[0] if len(gstin_matches) > 0 else ""
#         merchant_gstin = gstin_matches[1] if len(gstin_matches) > 1 else ""

#         hsn_code = extract(r"HSN Code:\s*(\d+)", text)
#         service_desc = extract(
#             r"Service Description:\s*(.*?)\s*Work Description:", text
#         )

#         sgst_rate = to_float(extract(r"SGST\s*@\s*([0-9.]+)%", text))
#         cgst_rate = to_float(extract(r"CGST\s*@\s*([0-9.]+)%", text))
#         igst_rate = to_float(extract(r"IGST\s*@\s*([0-9.]+)%", text))

#         invoice_total_amount = to_float(
#             extract(r"Total Amount\s*([\d,]+\.\d+)", text)
#         )

#         matches = re.findall(r"\d+\s+(.*?)\s+([\d,]+\.\d+)", text)

#         rows = []
#         ignore_words = [
#             "Taxable Amount",
#             "SGST",
#             "CGST",
#             "IGST",
#             "Total Amount",
#         ]

#         for match in matches:
#             particular_name = re.sub(r"^\d+\s*", "", match[0].strip())

#             if any(
#                 word.lower() in particular_name.lower() for word in ignore_words
#             ):
#                 continue

#             taxable_value = to_float(match[1])

#             sgst_amount = round(taxable_value * sgst_rate / 100, 2)
#             cgst_amount = round(taxable_value * cgst_rate / 100, 2)
#             igst_amount = round(taxable_value * igst_rate / 100, 2)

#             total_amount = round(
#                 taxable_value + sgst_amount + cgst_amount + igst_amount, 2
#             )

#             rows.append({
#                 "File Name": os.path.basename(pdf_file),
#                 "Invoice No": invoice_no,
#                 "Invoice Date": invoice_date,
#                 "Invoice Period": invoice_period,
#                 "Merchant Name": merchant_name,
#                 "Merchant ID": merchant_id,
#                 "Restaurant ID": restaurant_id,
#                 "Company GSTIN": company_gstin,
#                 "Merchant GSTIN": merchant_gstin,
#                 "HSN Code": hsn_code,
#                 "Service Description": service_desc,
#                 "Particular": particular_name,
#                 "Taxable Value": taxable_value,
#                 "SGST Rate (%)": sgst_rate,
#                 "SGST Amount": sgst_amount,
#                 "CGST Rate (%)": cgst_rate,
#                 "CGST Amount": cgst_amount,
#                 "IGST Rate (%)": igst_rate,
#                 "IGST Amount": igst_amount,
#                 "Total Amount": total_amount,
#                 "Invoice Total Amount": invoice_total_amount,
#             })

#         return rows

#     except Exception as e:
#         print(f"Error processing file {pdf_file}: {str(e)}")
#         return []


# # =====================================================
# # MAIN UPLOAD & AUTOMATIC DOWNLOAD API WITH TEMPFILE
# # =====================================================
# async def upload_zomato(
#     files: List[UploadFile] = File(...),
#     current_user=Depends(get_current_user),
#     db: Session = Depends(get_db),
# ):
#     all_data = []
#     files_count = len(files)

#     # 1. Incoming PDFs ko uploads directory me write karein
#     for file in files:
#         safe_filename = (
#             os.path.basename(file.filename).replace("/", "_").replace("\\", "_")
#         )
#         save_path = os.path.join(UPLOAD_FOLDER, safe_filename)

#         with open(save_path, "wb") as buffer:
#             shutil.copyfileobj(file.file, buffer)

#         all_data.extend(process_pdf(save_path))

#     # 2. DataFrame validation
#     df = pd.DataFrame(all_data)
#     if df.empty:
#         df = pd.DataFrame([{"Message": "No Data Found"}])

#     # 3. FIX: Create NamedTemporaryFile for Excel Generation
#     temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")

#     # 4. Excel write inside temporary path
#     with pd.ExcelWriter(temp_file.name, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False, sheet_name="Zomato_Invoices")

#     # 5. OpenPyXL auto-fit width changes over temporary workbook
#     wb = load_workbook(temp_file.name)
#     ws = wb.active

#     for column_cells in ws.columns:
#         length = max(
#             len(str(cell.value)) if cell.value else 0 for cell in column_cells
#         )
#         ws.column_dimensions[column_cells[0].column_letter].width = length + 5

#     wb.save(temp_file.name)
#     wb.close()

#     print(f"Temporary Zomato Excel written to: {temp_file.name}")

#     # 6. Database Analytics Tracking log update
#     if files_count > 0:
#         analytics = UploadAnalytics(
#             user_id=current_user.id, source="zomato", pdf_count=files_count
#         )
#         db.add(analytics)
#         db.commit()

#     # 7. CHROME AUTOMATIC DOWNLOAD WITH TEMPFILE & EXPOSED HEADERS
#     return FileResponse(
#         path=temp_file.name,
#         filename="Zomato_Invoice_Output.xlsx",
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Content-Disposition": 'attachment; filename="Zomato_Invoice_Output.xlsx"',
#             "Access-Control-Expose-Headers": "Content-Disposition",
#         },
#     )



# ----------------------------------------------------------------------------------------------------------------------

# import os
# import re
# import time
# import tempfile
# import shutil
# from datetime import datetime
# from concurrent.futures import ThreadPoolExecutor
# from typing import List

# import pandas as pd
# import pdfplumber
# from fastapi import Depends, File, UploadFile, BackgroundTasks, HTTPException
# from fastapi.responses import FileResponse
# from openpyxl import load_workbook
# from sqlalchemy.orm import Session

# # Database and Auth modules
# from auth.dependencies import get_current_user
# from database.database import get_db
# from database.models import UploadAnalytics


# # =====================================================
# # HELPER FUNCTIONS
# # =====================================================
# def extract(pattern, text):
#     match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
#     return match.group(1).strip() if match else ""


# def to_float(value):
#     if not value:
#         return 0.0
#     try:
#         return float(value.replace(",", "").strip())
#     except:
#         return 0.0


# def remove_file(path: str):
#     """Callback function to safely delete temporary Excel sheets after download"""
#     time.sleep(3)  # Wait for file stream to safely release
#     try:
#         if os.path.exists(path):
#             os.remove(path)
#     except Exception:
#         pass


# # =====================================================
# # CORE PDF PARSER FUNCTION
# # =====================================================
# def process_pdf(pdf_path_or_stream):
#     try:
#         base_name = os.path.basename(pdf_path_or_stream) if isinstance(pdf_path_or_stream, str) else "Zomato_Invoice.pdf"
        
#         with pdfplumber.open(pdf_path_or_stream) as pdf:
#             if not pdf.pages:
#                 return []
#             text = pdf.pages[0].extract_text() or ""

#         invoice_no = extract(r"Invoice No:\s*([A-Z0-9\-]+)", text)
#         invoice_date = extract(r"Invoice Date:\s*([0-9\-]+)", text)
#         invoice_period = extract(r"Period:\s*(.*?)\n", text)
#         merchant_name = extract(r"Entity Name:\s*(.*?)\s*State:", text)
#         merchant_id = extract(r"Merchant ID:\s*(\d+)", text)
#         restaurant_id = extract(r"Restaurant ID:\s*(\d+)", text)

#         gstin_matches = re.findall(r"GSTIN:\s*([0-9A-Z]{15})", text)
#         company_gstin = gstin_matches[0] if len(gstin_matches) > 0 else ""
#         merchant_gstin = gstin_matches[1] if len(gstin_matches) > 1 else ""

#         hsn_code = extract(r"HSN Code:\s*(\d+)", text)
#         service_desc = extract(r"Service Description:\s*(.*?)\s*Work Description:", text)

#         sgst_rate = to_float(extract(r"SGST\s*@\s*([0-9.]+)%", text))
#         cgst_rate = to_float(extract(r"CGST\s*@\s*([0-9.]+)%", text))
#         igst_rate = to_float(extract(r"IGST\s*@\s*([0-9.]+)%", text))

#         invoice_total_amount = to_float(extract(r"Total Amount\s*([\d,]+\.\d+)", text))

#         matches = re.findall(r"\d+\s+(.*?)\s+([\d,]+\.\d+)", text)

#         rows = []
#         ignore_words = ["Taxable Amount", "SGST", "CGST", "IGST", "Total Amount"]

#         for match in matches:
#             particular_name = re.sub(r"^\d+\s*", "", match[0].strip())
#             if any(word.lower() in particular_name.lower() for word in ignore_words):
#                 continue

#             taxable_value = to_float(match[1])
#             sgst_amount = round(taxable_value * sgst_rate / 100, 2)
#             cgst_amount = round(taxable_value * cgst_rate / 100, 2)
#             igst_amount = round(taxable_value * igst_rate / 100, 2)
#             total_amount = round(taxable_value + sgst_amount + cgst_amount + igst_amount, 2)

#             rows.append({
#                 "File Name": base_name,
#                 "Invoice No": invoice_no,
#                 "Invoice Date": invoice_date,
#                 "Invoice Period": invoice_period,
#                 "Merchant Name": merchant_name,
#                 "Merchant ID": merchant_id,
#                 "Restaurant ID": restaurant_id,
#                 "Company GSTIN": company_gstin,
#                 "Merchant GSTIN": merchant_gstin,
#                 "HSN Code": hsn_code,
#                 "Service Description": service_desc,
#                 "Particular": particular_name,
#                 "Taxable Value": taxable_value,
#                 "SGST Rate (%)": sgst_rate,
#                 "SGST Amount": sgst_amount,
#                 "CGST Rate (%)": cgst_rate,
#                 "CGST Amount": cgst_amount,
#                 "IGST Rate (%)": igst_rate,
#                 "IGST Amount": igst_amount,
#                 "Total Amount": total_amount,
#                 "Invoice Total Amount": invoice_total_amount,
#             })

#         return rows
#     except Exception as e:
#         print(f"Error processing file: {str(e)}")
#         return []


# # =====================================================
# # RENDER TIERS RAM OPTIMIZED API
# # =====================================================
# async def upload_zomato(
#     background_tasks: BackgroundTasks,
#     files: List[UploadFile],
#     db: Session,
#     current_user
# ):
#     all_data = []
#     files_count = len(files)
#     temp_pdf_paths = []

#     try:
#         # CHUNKS READ RUNNER: RAM spikes ko completely rokne ke liye
#         for file in files:
#             tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
#             try:
#                 shutil.copyfileobj(file.file, tmp)
#             finally:
#                 tmp.close()
#                 file.file.close()
#             temp_pdf_paths.append(tmp.name)

#         # Render Core optimization: Max workers ko 4 kiya taaki CPU halt na ho
#         with ThreadPoolExecutor(max_workers=4) as executor:
#             results = list(executor.map(process_pdf, temp_pdf_paths))

#         for row_list in results:
#             if row_list:
#                 all_data.extend(row_list)

#         df = pd.DataFrame(all_data)
#         if df.empty:
#             df = pd.DataFrame([{"Message": "No Data Found"}])

#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#         excel_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
#         excel_filename = f"Zomato_Invoice_Output_{timestamp}.xlsx"

#         with pd.ExcelWriter(excel_temp.name, engine="openpyxl") as writer:
#             df.to_excel(writer, index=False, sheet_name="Zomato_Invoices")

#         wb = load_workbook(excel_temp.name)
#         ws = wb.active

#         for column_cells in ws.columns:
#             length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
#             ws.column_dimensions[column_cells[0].column_letter].width = length + 5

#         wb.save(excel_temp.name)
#         wb.close()

#         if files_count > 0 and current_user:
#             analytics = UploadAnalytics(
#                 user_id=current_user.id, source="zomato", pdf_count=files_count
#             )
#             db.add(analytics)
#             db.commit()

#         background_tasks.add_task(remove_file, excel_temp.name)

#         return FileResponse(
#             path=excel_temp.name,
#             filename=excel_filename,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={
#                 "Content-Disposition": f'attachment; filename="{excel_filename}"',
#                 "Access-Control-Expose-Headers": "Content-Disposition"
#             }
#         )

#     except Exception as e:
#         print(f"Global Exception Handler Caught: {str(e)}")
#         raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

#     finally:
#         for path in temp_pdf_paths:
#             try:
#                 if os.path.exists(path):
#                     os.remove(path)
#             except:
#                 pass


# ------------------------------------------------------------------------------------------------------------




import os
import re
import time
import tempfile
import shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from typing import List

import pandas as pd
import pdfplumber
from fastapi import Depends, File, UploadFile, BackgroundTasks, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

# Database modules
from database.database import get_db
from database.models import UploadAnalytics


# =====================================================
# HELPER FUNCTIONS
# =====================================================
def extract(pattern, text):
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return match.group(1).strip() if match else ""


def to_float(value):
    if not value:
        return 0.0
    try:
        return float(value.replace(",", "").strip())
    except:
        return 0.0


def remove_file(path: str):
    """Callback function to safely delete temporary Excel sheets after download"""
    time.sleep(3)  # Wait for file stream to safely release
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass


# =====================================================
# CORE PDF PARSER FUNCTION
# =====================================================
def process_pdf(pdf_path_or_stream):
    try:
        base_name = os.path.basename(pdf_path_or_stream) if isinstance(pdf_path_or_stream, str) else "Zomato_Invoice.pdf"
        
        with pdfplumber.open(pdf_path_or_stream) as pdf:
            if not pdf.pages:
                return []
            text = pdf.pages[0].extract_text() or ""

        invoice_no = extract(r"Invoice No:\s*([A-Z0-9\-]+)", text)
        invoice_date = extract(r"Invoice Date:\s*([0-9\-]+)", text)
        invoice_period = extract(r"Period:\s*(.*?)\n", text)
        merchant_name = extract(r"Entity Name:\s*(.*?)\s*State:", text)
        merchant_id = extract(r"Merchant ID:\s*(\d+)", text)
        restaurant_id = extract(r"Restaurant ID:\s*(\d+)", text)

        gstin_matches = re.findall(r"GSTIN:\s*([0-9A-Z]{15})", text)
        company_gstin = gstin_matches[0] if len(gstin_matches) > 0 else ""
        merchant_gstin = gstin_matches[1] if len(gstin_matches) > 1 else ""

        hsn_code = extract(r"HSN Code:\s*(\d+)", text)
        service_desc = extract(r"Service Description:\s*(.*?)\s*Work Description:", text)

        sgst_rate = to_float(extract(r"SGST\s*@\s*([0-9.]+)%", text))
        cgst_rate = to_float(extract(r"CGST\s*@\s*([0-9.]+)%", text))
        igst_rate = to_float(extract(r"IGST\s*@\s*([0-9.]+)%", text))

        invoice_total_amount = to_float(extract(r"Total Amount\s*([\d,]+\.\d+)", text))

        matches = re.findall(r"\d+\s+(.*?)\s+([\d,]+\.\d+)", text)

        rows = []
        ignore_words = ["Taxable Amount", "SGST", "CGST", "IGST", "Total Amount"]

        for match in matches:
            particular_name = re.sub(r"^\d+\s*", "", match[0].strip())
            if any(word.lower() in particular_name.lower() for word in ignore_words):
                continue

            taxable_value = to_float(match[1])
            sgst_amount = round(taxable_value * sgst_rate / 100, 2)
            cgst_amount = round(taxable_value * cgst_rate / 100, 2)
            igst_amount = round(taxable_value * igst_rate / 100, 2)
            total_amount = round(taxable_value + sgst_amount + cgst_amount + igst_amount, 2)

            rows.append({
                "File Name": base_name,
                "Invoice No": invoice_no,
                "Invoice Date": invoice_date,
                "Invoice Period": invoice_period,
                "Merchant Name": merchant_name,
                "Merchant ID": merchant_id,
                "Restaurant ID": restaurant_id,
                "Company GSTIN": company_gstin,
                "Merchant GSTIN": merchant_gstin,
                "HSN Code": hsn_code,
                "Service Description": service_desc,
                "Particular": particular_name,
                "Taxable Value": taxable_value,
                "SGST Rate (%)": sgst_rate,
                "SGST Amount": sgst_amount,
                "CGST Rate (%)": cgst_rate,
                "CGST Amount": cgst_amount,
                "IGST Rate (%)": igst_rate,
                "IGST Amount": igst_amount,
                "Total Amount": total_amount,
                "Invoice Total Amount": invoice_total_amount,
            })

        return rows
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return []


# =====================================================
# RENDER TIERS RAM OPTIMIZED API
# =====================================================
async def upload_zomato(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user = None
):
    all_data = []
    files_count = len(files)
    temp_pdf_paths = []

    try:
        # CHUNKS READ RUNNER: Prevents RAM spikes
        for file in files:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            try:
                shutil.copyfileobj(file.file, tmp)
            finally:
                tmp.close()
                file.file.close()
            temp_pdf_paths.append(tmp.name)

        # ThreadPool optimization for CPU utilization
        with ThreadPoolExecutor(max_workers=4) as executor:
            results = list(executor.map(process_pdf, temp_pdf_paths))

        for row_list in results:
            if row_list:
                all_data.extend(row_list)

        df = pd.DataFrame(all_data)
        if df.empty:
            df = pd.DataFrame([{"Message": "No Data Found"}])

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_filename = f"Zomato_Invoice_Output_{timestamp}.xlsx"

        with pd.ExcelWriter(excel_temp.name, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Zomato_Invoices")

        wb = load_workbook(excel_temp.name)
        ws = wb.active

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        wb.save(excel_temp.name)
        wb.close()

        if files_count > 0 and current_user and hasattr(current_user, 'id'):
            analytics = UploadAnalytics(
                user_id=current_user.id,
                source="zomato",
                pdf_count=files_count
            )
            db.add(analytics)
            db.commit()

        background_tasks.add_task(remove_file, excel_temp.name)

        return FileResponse(
            path=excel_temp.name,
            filename=excel_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{excel_filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )

    except Exception as e:
        print(f"Global Exception Handler Caught: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

    finally:
        for path in temp_pdf_paths:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass